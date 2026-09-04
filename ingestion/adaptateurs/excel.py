"""Adaptateur generique -- fichier Excel. Un seul lecteur reutilisable,
pas un par domaine. `data_only=False` : garde les formules cassees
(#REF!, etc.) visibles comme texte plutot que de les faire disparaitre --
c'est justement ce genre de defaut qu'on veut voir arriver en brut."""

from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook


def lire_excel(
    path: Path, sheet_name: str | None = None, header_row: int = 1
) -> list[dict[str, object]]:
    wb = load_workbook(path, data_only=False)
    ws = wb[sheet_name] if sheet_name else wb.active
    rows = list(ws.iter_rows(min_row=header_row, values_only=True))
    if not rows:
        return []
    headers = [str(h) if h is not None else f"col{i}" for i, h in enumerate(rows[0])]
    return [dict(zip(headers, row)) for row in rows[1:] if any(v is not None for v in row)]
