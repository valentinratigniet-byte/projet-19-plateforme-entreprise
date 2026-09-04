"""Simule la grille de remises negociees, tenue a la main par l'equipe
commerciale en parallele de l'AS/400 -- le cas classique du "shadow IT"
qu'aucune base interne ne couvre bien.

Genere DEUX fichiers, comme deux commerciaux qui tiennent chacun leur
propre suivi sans synchronisation reelle :
  remises_ventes_v3.xlsx        (ancienne version, encore utilisee par certains)
  remises_ventes_v4_FINAL.xlsx  (plus recente, mais pas totalement alignee)

Defauts reels injectes :
  - cellules fusionnees (titre)
  - clients references par NOM saisi a la main (pas par CLICOD), avec
    variantes de casse/orthographe -- ne matche pas exactement les noms
    AS/400
  - formule cassee (#REF!, cellule source supprimee) sur une ligne
  - valeurs de remise qui DIVERGENT entre les deux fichiers pour le meme
    client -- vrai conflit a trancher lors du nettoyage, pas invente
"""

from __future__ import annotations

import random
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font

OUT_DIR = Path(__file__).parent / "exports"

# Sous-ensemble de noms de clients "tels que tapes a la main" -- variantes
# volontaires par rapport aux CLINOM generes par le simulateur AS/400
# (rapprochement flou a faire plus tard, meme discipline que le Projet 12).
CLIENTS_SAISIS = [
    "Dubois & Fils",
    "DUBOIS ET FILS",
    "Martin Distribution",
    "Martin Distrib.",
    "Lefevre SARL",
    "Lefèvre Sarl",
    "Groupe Bernard",
    "Bernard Groupe",
    "Petit Commerce",
    "Moreau Négoce",
    "Moreau Negoce",
    "Girard Industries",
    "Girard Ind.",
    "Roux Distribution",
    "Fournier & Cie",
    "Fournier et Cie",
]


def construire_v3(rng: random.Random) -> dict[str, float]:
    wb = Workbook()
    ws = wb.active
    ws.title = "Remises"

    ws.merge_cells("A1:D1")
    ws["A1"] = "Grille de remises negociees -- V3 (a jour au 15/02/2026)"
    ws["A1"].font = Font(bold=True, size=14)

    ws.append(["Client", "Remise (%)", "Commentaire", "Valable jusqu'au"])

    remises = {}
    for i, client in enumerate(CLIENTS_SAISIS[::2], start=3):  # 1 nom sur 2
        remise = rng.choice([5, 8, 10, 12, 15])
        remises[client] = remise
        ws.cell(row=i, column=1, value=client)
        ws.cell(row=i, column=2, value=remise)
        ws.cell(row=i, column=3, value="Negocie par commercial" if rng.random() < 0.5 else "")
        ws.cell(row=i, column=4, value="31/12/2026")

    # Formule cassee : reference une feuille/cellule qui n'existe plus
    row_cassee = ws.max_row + 1
    ws.cell(row=row_cassee, column=1, value="Ancien Client SA")
    ws.cell(row=row_cassee, column=2, value="=Grille2026!B99")  # #REF! a l'ouverture
    ws.cell(row=row_cassee, column=3, value="a verifier - feuille supprimee")

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_DIR / "remises_ventes_v3.xlsx")
    return remises


def construire_v4(rng: random.Random, remises_v3: dict[str, float]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Remises"

    ws.merge_cells("A1:D1")
    ws["A1"] = "Grille de remises negociees -- V4_FINAL (a jour au 28/07/2026)"
    ws["A1"].font = Font(bold=True, size=14)

    ws.append(["Client", "Remise (%)", "Commentaire", "Valable jusqu'au"])

    row = 3
    for client in CLIENTS_SAISIS:  # tous les noms cette fois, avec variantes de casse
        if client in remises_v3 and rng.random() < 0.4:
            # ~40% des remises reprises de la V3 ont ete modifiees depuis --
            # divergence reelle entre les 2 fichiers, pas juste une copie
            remise = remises_v3[client] + rng.choice([-3, -2, 2, 3, 5])
        elif client in remises_v3:
            remise = remises_v3[client]  # inchangee
        else:
            remise = rng.choice([5, 8, 10, 12, 15, 20])

        ws.cell(row=row, column=1, value=client)
        ws.cell(row=row, column=2, value=remise)
        ws.cell(row=row, column=3, value="")
        ws.cell(row=row, column=4, value="31/12/2026" if rng.random() < 0.8 else "")
        row += 1

    OUT_DIR.mkdir(parents=True, exist_ok=True)
    wb.save(OUT_DIR / "remises_ventes_v4_FINAL.xlsx")


def main() -> None:
    rng = random.Random(19)
    remises_v3 = construire_v3(rng)
    construire_v4(rng, remises_v3)
    print(f"2 fichiers generes dans {OUT_DIR}")


def _self_check() -> None:
    from openpyxl import load_workbook

    for name in ("remises_ventes_v3.xlsx", "remises_ventes_v4_FINAL.xlsx"):
        path = OUT_DIR / name
        assert path.exists(), f"{name} manquant"
        wb = load_workbook(path)
        assert wb.active.max_row > 1, f"{name}: pas de donnees"
    print("self-check OK: les 2 fichiers de remises existent et contiennent des lignes")


if __name__ == "__main__":
    main()
    _self_check()
